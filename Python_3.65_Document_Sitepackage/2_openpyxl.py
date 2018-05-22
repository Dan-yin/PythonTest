#import openpyxl
from openpyxl import Workbook
import datetime



def  test():
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 420
    ws.append([1,2,3,4,5])
    #ws.column_dimensions = 21
    ws['A2'] = datetime.datetime.now()
    ws1 = wb.create_sheet(title = 'Mysheet')
    ws2 = wb.create_sheet(title = 'MySheet',index = 0)
    
    ws.title = 'new Title'
    ws.sheet_properties.tabColor = '1072BA'
    ws3 = wb['new Title']
    #print(ws3 is ws) # True
    #print(wb.sheetnames)

    for sheet in wb:
        print(sheet.title)

    target = wb.copy_worksheet(ws).title = 'this'

    c = ws['A4']
    ws['A4']  = 4
    
    ws.append(range(0,10))
        
    def save(wb):    
        for i in range(10):
            try :
                wb.save('sample{0}.xlsx'.format(str(i)))
                break
            except PermissionError:
                pass

    save(wb)

if __name__ == '__main__':
    test()